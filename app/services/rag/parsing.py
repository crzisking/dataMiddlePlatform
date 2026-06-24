"""文档解析：从不同格式的文件里把"纯文字"抽取出来。

为什么要解析：RAG 只能处理文字。PDF/Word/Excel 各有各的二进制格式，
得用对应的库把里面的文字读出来，后面才能切割、向量化。

当前支持：PDF(原生文字) / Word(docx) / Excel(xlsx) / 纯文本(txt、md)。
还没做：扫描件 PDF 的 OCR（图片里的字），老格式 .doc/.xls。
"""

import io

import fitz  # PyMuPDF：读 PDF 的库
from docx import Document as DocxDocument  # 读 Word(docx) 的库
from openpyxl import load_workbook  # 读 Excel(xlsx) 的库

from app.core.exceptions import BadRequestError


def extract_text(ext: str, data: bytes) -> str:
    """根据文件扩展名，选对应的解析方式，返回抽取出的纯文字。"""
    if ext == "pdf":
        return _pdf(data)
    if ext == "docx":
        return _docx(data)
    if ext == "xlsx":
        return _xlsx(data)
    if ext in ("txt", "md"):
        # 纯文本直接按 UTF-8 解码；遇到无法解码的字节就忽略，不让整篇失败
        return data.decode("utf-8", errors="ignore")
    if ext in ("doc", "xls"):
        # 老二进制格式解析麻烦，暂不支持，直接报错让用户转存成新格式
        raise BadRequestError(f"老格式 .{ext} 暂不支持解析，请另存为 .docx/.xlsx 再上传")
    raise BadRequestError(f"不支持解析的格式 .{ext}")


def _pdf(data: bytes) -> str:
    # 从内存里的字节打开 PDF（不落磁盘），逐页取文字后用换行拼起来
    with fitz.open(stream=data, filetype="pdf") as doc:
        return "\n".join(page.get_text() for page in doc)


def _docx(data: bytes) -> str:
    # docx 由一段段段落组成，逐段取文字；跳过空段落
    d = DocxDocument(io.BytesIO(data))
    return "\n".join(p.text for p in d.paragraphs if p.text.strip())


def _xlsx(data: bytes) -> str:
    # data_only=True：取单元格"算出来的值"而不是公式本身
    # read_only=True：只读模式，省内存（大表格友好）
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        lines.append(f"# 工作表：{ws.title}")
        for row in ws.iter_rows(values_only=True):
            # 一行里非空的单元格用 " | " 连起来，让整行作为一条，保留表格的横向语义
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(" | ".join(cells))
    wb.close()
    return "\n".join(lines)
